import logging
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config.settings import OUTPUT_DIR, PRODUCT_SCHEMA
from src.file_manager import FileManager


class DataProcessor:
    """Processeur de données pour transformer les résultats JSON en fichiers exploitables"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.file_manager = FileManager()
        self.output_dir = Path(OUTPUT_DIR)
        
    def process_results(self, raw_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Traite et nettoie les résultats bruts de Gemini
        
        Args:
            raw_results (List[Dict]): Résultats bruts de l'analyse
            
        Returns:
            List[Dict]: Données nettoyées et validées
        """
        self.logger.info(f"📊 Traitement de {len(raw_results)} résultats")
        
        processed_data = []
        stats = {
            'total': len(raw_results),
            'valides': 0,
            'erreurs': 0,
            'partiels': 0
        }
        
        for i, result in enumerate(raw_results, 1):
            try:
                # Nettoyer et valider les données
                cleaned_result = self._clean_result(result)
                
                # Ajouter des métadonnées
                cleaned_result['numero_sequence'] = i
                cleaned_result['date_traitement'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Calculer le score de complétude
                completeness_score = self._calculate_completeness(cleaned_result)
                cleaned_result['score_completude'] = completeness_score
                
                # Catégoriser le résultat
                if 'erreur' in cleaned_result:
                    stats['erreurs'] += 1
                    cleaned_result['statut'] = 'Erreur'
                elif completeness_score >= 80:
                    stats['valides'] += 1
                    cleaned_result['statut'] = 'Complet'
                else:
                    stats['partiels'] += 1
                    cleaned_result['statut'] = 'Partiel'
                
                processed_data.append(cleaned_result)
                
            except Exception as e:
                self.logger.error(f"Erreur lors du traitement du résultat {i}: {e}")
                stats['erreurs'] += 1
                continue
        
        self._log_processing_stats(stats)
        return processed_data
    
    def _clean_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Nettoie un résultat individuel
        
        Args:
            result (Dict): Résultat brut
            
        Returns:
            Dict: Résultat nettoyé
        """
        cleaned = {}
        
        # Nettoyer chaque champ
        for field in PRODUCT_SCHEMA.keys():
            value = result.get(field, "Non détecté")
            
            if isinstance(value, str):
                # Nettoyer les espaces et caractères spéciaux
                value = value.strip()
                
                # Traiter les cas spéciaux
                if value.lower() in ['', 'n/a', 'na', 'null', 'none', '?', '-']:
                    value = "Non détecté"
                elif field == 'prix_fcfa':
                    value = self._clean_price(value)
                elif field == 'volume':
                    value = self._clean_volume(value)
                elif field == 'code_barres_ean':
                    value = self._clean_barcode(value)
            
            cleaned[field] = value
        
        # Conserver les métadonnées
        for meta_field in ['nom_fichier', 'chemin_fichier', 'erreur']:
            if meta_field in result:
                cleaned[meta_field] = result[meta_field]
        
        return cleaned
    
    def _clean_price(self, price_str: str) -> str:
        """Nettoie et standardise les prix"""
        if price_str == "Non détecté":
            return price_str
        
        # Supprimer les caractères non numériques sauf points et virgules
        import re
        cleaned = re.sub(r'[^\d.,]', '', price_str)
        
        if not cleaned:
            return "Non détecté"
        
        return cleaned
    
    def _clean_volume(self, volume_str: str) -> str:
        """Nettoie et standardise les volumes"""
        if volume_str == "Non détecté":
            return volume_str
        
        # Standardiser les unités communes
        volume_str = volume_str.lower()
        replacements = {
            'litres': 'L',
            'litre': 'L',
            'millilitres': 'mL',
            'millilitre': 'mL',
            'grammes': 'g',
            'gramme': 'g',
            'kilogrammes': 'kg',
            'kilogramme': 'kg'
        }
        
        for old, new in replacements.items():
            volume_str = volume_str.replace(old, new)
        
        return volume_str.strip()
    
    def _clean_barcode(self, barcode_str: str) -> str:
        """Nettoie les codes-barres"""
        if barcode_str == "Non détecté":
            return barcode_str
        
        # Garder seulement les chiffres
        import re
        cleaned = re.sub(r'[^\d]', '', barcode_str)
        
        if not cleaned:
            return "Non détecté"
        
        return cleaned
    
    def _calculate_completeness(self, result: Dict[str, Any]) -> float:
        """
        Calcule le score de complétude d'un résultat
        
        Args:
            result (Dict): Résultat à évaluer
            
        Returns:
            float: Score de complétude (0-100)
        """
        total_fields = len(PRODUCT_SCHEMA)
        complete_fields = 0
        
        for field in PRODUCT_SCHEMA.keys():
            value = result.get(field, "")
            if value and value != "Non détecté" and value != "Erreur d'analyse":
                complete_fields += 1
        
        return (complete_fields / total_fields) * 100
    
    def generate_output_files(self, processed_data: List[Dict[str, Any]]) -> Dict[str, Path]:
        """
        Génère les fichiers de sortie (CSV et Excel)
        
        Args:
            processed_data (List[Dict]): Données traitées
            
        Returns:
            Dict[str, Path]: Chemins des fichiers générés
        """
        session_id = self.file_manager.create_session_id()
        
        # Générer les noms de fichiers
        csv_filename = f"resultats_analyse_{session_id}.csv"
        excel_filename = f"resultats_analyse_{session_id}.xlsx"
        
        csv_path = self.output_dir / csv_filename
        excel_path = self.output_dir / excel_filename
        
        self.logger.info(f"📁 Génération des fichiers de sortie")
        
        # Générer le CSV
        self._generate_csv(processed_data, csv_path)
        
        # Générer l'Excel avec mise en forme
        self._generate_excel(processed_data, excel_path)
        
        return {
            'csv': csv_path,
            'excel': excel_path
        }
    
    def _generate_csv(self, data: List[Dict[str, Any]], output_path: Path):
        """Génère le fichier CSV"""
        try:
            df = pd.DataFrame(data)
            
            # Réorganiser les colonnes dans un ordre logique
            column_order = [
                'numero_sequence', 'nom_fichier', 'nom_produit', 'description_type',
                'volume', 'prix_fcfa', 'code_barres_ean', 'code_article',
                'source_information', 'statut', 'score_completude',
                'date_traitement', 'chemin_fichier'
            ]
            
            # Garder seulement les colonnes qui existent
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            self.logger.info(f"✅ Fichier CSV généré: {output_path.name}")
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération CSV: {e}")
            raise
    
    def _generate_excel(self, data: List[Dict[str, Any]], output_path: Path):
        """Génère le fichier Excel avec mise en forme"""
        try:
            df = pd.DataFrame(data)
            
            # Réorganiser les colonnes
            column_order = [
                'numero_sequence', 'nom_fichier', 'nom_produit', 'description_type',
                'volume', 'prix_fcfa', 'code_barres_ean', 'code_article',
                'source_information', 'statut', 'score_completude',
                'date_traitement'
            ]
            
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            # Sauvegarder avec openpyxl pour la mise en forme
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Résultats', index=False)
                
                # Mise en forme
                worksheet = writer.sheets['Résultats']
                self._format_excel_worksheet(worksheet, df)
            
            self.logger.info(f"✅ Fichier Excel généré: {output_path.name}")
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération Excel: {e}")
            raise
    
    def _format_excel_worksheet(self, worksheet, df):
        """Applique la mise en forme au fichier Excel"""
        # En-têtes en gras avec fond coloré
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _log_processing_stats(self, stats: Dict[str, int]):
        """Log les statistiques de traitement"""
        self.logger.info("📈 Statistiques de traitement:")
        self.logger.info(f"   Total: {stats['total']}")
        self.logger.info(f"   Complets: {stats['valides']}")
        self.logger.info(f"   Partiels: {stats['partiels']}")
        self.logger.info(f"   Erreurs: {stats['erreurs']}")
        
        if stats['total'] > 0:
            success_rate = ((stats['valides'] + stats['partiels']) / stats['total']) * 100
            self.logger.info(f"   Taux de succès: {success_rate:.1f}%")
    
    def generate_summary_report(self, processed_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Génère un rapport de synthèse
        
        Args:
            processed_data (List[Dict]): Données traitées
            
        Returns:
            Dict: Rapport de synthèse
        """
        total_images = len(processed_data)
        if total_images == 0:
            return {"error": "Aucune donnée à analyser"}
        
        # Calculer les statistiques
        statuts = [item.get('statut', 'Inconnu') for item in processed_data]
        scores = [item.get('score_completude', 0) for item in processed_data if isinstance(item.get('score_completude'), (int, float))]
        
        report = {
            'total_images': total_images,
            'images_completes': statuts.count('Complet'),
            'images_partielles': statuts.count('Partiel'),
            'images_erreur': statuts.count('Erreur'),
            'score_moyen': sum(scores) / len(scores) if scores else 0,
            'taux_succes': ((statuts.count('Complet') + statuts.count('Partiel')) / total_images) * 100,
            'date_traitement': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        return report